# -*- coding: utf-8 -*-
"""
Created on Tue Jan 12 12:11:54 2016

@author: RBdC
"""
from casadi import *
from casadi.tools import *
from matplotlib import pylab as plt
import math
import scipy.linalg as scla
import numpy as np
from Utilities import*
#
import control.matlab as cnt
#
import xlrd
#import xlsxwriter
import openpyxl

### 1) Simulation Fundamentals

# 1.1) Simulation discretization parameters
Nsim = 500  # Simulation length

N = 50      # Horizon

h = 1       # Time step 

# Identified TF system from SIPPY (2x1 ARX system)
den = [1.0, -0.9333]
# den = den1 = den2
# num1: 1->1 
num1 = [-0.3273]
# num1: 2->1 
num2 = [0.029];
#
NUM = [[num1,num2]]
DEN = [[den,den]]
#
G1 = cnt.tf(num1, den,h)
G2 = cnt.tf(NUM, DEN,h)
# SS system
Gss1 = cnt.tf2ss(G1)        # 1x1
Gss2 = cnt.tf2ss(G2)        # 2x1

# 3.1.2) Symbolic variables
xp = SX.sym("xp", 1)    # process state vector        
x = SX.sym("x", 1)      # model state vector          
u = SX.sym("u", 1)      # control vector              
y = SX.sym("y", 1)      # measured output vector      
d = SX.sym("d", 1)      # disturbance                 

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
### 2) Process and Model construction 

# Model =/= Process ?? maybe not if dis is measured

# 2.1) Process Parameters
Ap = np.array(Gss2.A)
Bp = np.array(Gss2.B[0,0])
Cp = np.array(Gss2.C)


# Include real disturbance

# Import Real Industrial Data
ExcelSheet = 'Dati_test_GBN_aprile_2021.xlsx'

## Loading the file:
Book = xlrd.open_workbook(ExcelSheet)
# first sheet
first_sheet = Book.sheet_by_index(0)
# 2021 data set (complete)
Ndata = 5759    

# PROCESS
# calce idrata (spenta) - CaOH2
# comando coclea - coclea A
CocleaA = np.zeros((Ndata,1))
# portata calce - coclea A
Q2A_calce = np.zeros((Ndata,1))
# Conc. HCl 
HCl_smp1 = np.zeros((Ndata,1))   # input stadio 1 (@SMP1)
HCl_smp2 = np.zeros((Ndata,1))   # output stadio 1 (@SMP2)
#
# altro
SO2_smp1 = np.zeros((Ndata,1))   # input stadio 1 (@SMP1)
HF_smp1 = np.zeros((Ndata,1))   # input stadio 1 (@SMP1)
#
H2O_smp1 = np.zeros((Ndata,1))   # input stadio 1 (@SMP1)
H2O_smp2 = np.zeros((Ndata,1))   # input stadio 2 (@SMP1)
H2O_stack = np.zeros((Ndata,1))   # input camino (@SME)
#
O2_smp1 = np.zeros((Ndata,1))   # input stadio 1 (@SMP1)
O2_smp2 = np.zeros((Ndata,1))   # input stadio 2 (@SMP2)
O2_stack = np.zeros((Ndata,1))   # input camino (@SME)
#
Q_stack = np.zeros((Ndata,1))   # input camino (@SME)
Q_ryc = np.zeros((Ndata,1))     # input camino (@SME)

 

# read data: input-output
for i in range(Ndata):
    #
    CocleaA[i,:] = first_sheet.cell(4+i,1).value        # colonna B
    #
    Q2A_calce[i,:] = first_sheet.cell(4+i,3).value      # colonna D
    #
    HCl_smp1[i,:] = first_sheet.cell(4+i,15).value      # colonna P
    #
    HCl_smp2[i,:] = first_sheet.cell(4+i,24).value      # colonna Y
    # altro
    SO2_smp1[i,:] = first_sheet.cell(4+i,16).value      # colonna Q
    #HF_smp1[i,:] = first_sheet.cell(4+i,17).value      # colonna R
    #
    H2O_smp1[i,:] = first_sheet.cell(4+i,13).value      # colonna N
    O2_smp1[i,:] = first_sheet.cell(4+i,14).value      # colonna O
    #
    H2O_smp2[i,:] = first_sheet.cell(4+i,22).value      # colonna W
    O2_smp2[i,:] = first_sheet.cell(4+i,23).value      # colonna X
    #
    H2O_stack[i,:] = first_sheet.cell(4+i,30).value      # colonna AE
    O2_stack[i,:] = first_sheet.cell(4+i,31).value      # colonna AF
    #
    Q_stack[i,:] = first_sheet.cell(4+i,7).value      # colonna H
    Q_ryc[i,:] = first_sheet.cell(4+i,8).value      # colonna I
    

# define calibration data
Cal_value = 1600   
# connect gaps for calibration data
for i in range(Ndata):
    if HCl_smp1[i,:] == Cal_value:
        # set the previous value
        HCl_smp1[i,:] = HCl_smp1[i-1,:] 
             
##  Define clean data - validation set 
# (4 celle in testa --> -4)
t_in = 2070              # 
t_fin = 3906             #
# 
Qm_calce = np.mean(Q2A_calce)
HClm_smp1 = np.mean(HCl_smp1)
print("Qm_calce", Qm_calce)
print("HClm_smp1", HClm_smp1)

## Ouputs
Y_30 = HCl_smp2[t_in:t_fin+1,:].T
[p,N30] = Y_30.shape
Ym = np.mean(Y_30)
print("Ym", Ym);
    
## Inputs
U1 = Q2A_calce[t_in:t_fin+1,:].T
U1m = np.mean(U1)
print("U1m", U1m);
U2 = HCl_smp1[t_in:t_fin+1,:].T
U2m = np.mean(U2)
print("U2m", U2m);
U_30 = np.vstack((U1,U2))
    
## Resampling data: 30 --> 60 sec
# concentrazioni variano solamente ogni 60 secs
#
resampling = 'no'
if resampling == 'yes':
    Nd = int(np.floor(N30/2))
    U_data = np.zeros((2,Nd))
    Y_data = np.zeros((p,Nd))
    #     
    for i in range(Nd):
        U_data[:,i] = U_30[:,2*i]
        Y_data[:,i] = Y_30[:,2*i]
else:
    U_data = U_30.copy()
    Y_data = Y_30.copy() 
    # for i in range(N30):
    #     U_data[:,i] = U_30[:,i]
    #     Y_data[:,i] = Y_30[:,i]    
    
# Additive State Measurable Disturbances 
def defdxm(t):
    """
    SUMMARY:
    It constructs the meaurable linearly additive disturbances 
    
    SYNTAX:
    assignment = defdp(k)
  
    ARGUMENTS:
    + t             - Variable that indicate the current time
    
    OUTPUTS:
    + dxp      - State disturbance value
    """
    
    # pass external disturbance: Bd*d
    dxp = np.array([np.dot(Gss2.B[0,1],U_data[1,t])])
    
    
    return [dxp]

def defdxmp(t):
    """
    SUMMARY:
    It constructs the meaurable linearly additive disturbances 
    
    SYNTAX:
    assignment = defdp(k)
  
    ARGUMENTS:
    + t             - Variable that indicate the current time
    
    OUTPUTS:
    + dxp      - State disturbance value
    """
    
    # pass external disturbance: Bd*d
    dxp = 0.85*np.array([np.dot(Gss2.B[0,1],U_data[1,t])])
    
    
    return [dxp]

def defdxp(t):
    """
    SUMMARY:
    It constructs the meaurable linearly additive disturbances 
    
    SYNTAX:
    assignment = defdp(k)
  
    ARGUMENTS:
    + t             - Variable that indicate the current time
    
    OUTPUTS:
    + dxp      - State disturbance value
    """
    
    # pass external disturbance: Bd*d
    dxp = 0.15*np.array([np.dot(Gss2.B[0,1],U_data[1,t])])
    
    
    return [dxp]
# 2.2) Model Parameters
# to be changed here around after GP + MV meeting
MOD_type = 'extra'
if MOD_type == 'norm':
    A = np.array(Gss1.A)
    B = np.array(Gss1.B)
    C = np.array(Gss1.C)
elif MOD_type == 'extra':
    A = np.array(Gss2.A)
    B = np.array(Gss2.B[0,0])
    C = np.array(Gss2.C)
    

# 2.3) Disturbance model for Offset-free control
offree = "lin"
D_mod_type = 'IN'
if D_mod_type == 'OUT':
    # output DM
    Bd = np.zeros((x.size1(),d.size1()))
    Cd = np.eye(d.size1())
elif D_mod_type == 'IN':
    # input DM
    Bd = np.array(Gss2.B[0,1])
    Cd = np.zeros((y.size1(),d.size1()))
    

# 2.4) Initial condition
SP0 = 450
x0_p = SP0/Cp[0][0]*np.ones((xp.size1(),1))
x0_m = SP0/C[0][0]*np.ones((x.size1(),1))
#u0 = (1 - A[0][0])*x0_m/B[0]*np.ones((u.size1(),1))
u0 = (1 - A[0][0])*x0_m/B*np.ones((u.size1(),1))

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
### 3) State Estimation

# Dimensions 
nx = x.size1()  # state vector                #
nu = u.size1()  # control vector              #
ny = y.size1()  # measured output vector      #
nd = d.size1()  # disturbance                 #


# Estimator type
est_type = 'KF'

# Luemberger filter tuning params
if est_type == 'Lue':
    lue = True
    Kx = np.zeros((nx,nd))
    Kd = np.eye(nd)
    K = np.vstack((Kx,Kd))

# Kalman filter  
elif est_type == 'KF':
    kal = True # Set True if you want the Kalman filter
    Qx_kf = 1.0e-8*np.eye(nx)
    Qd_kf = np.eye(nd)
    Q_kf = scla.block_diag(Qx_kf, Qd_kf)
    R_kf = 1.0e-8*np.eye(ny)
    P0 = 1e-2*Q_kf
    # tipico tuning:
    # Qx + basso, filtro + aggressivo
    # P0 basso: non mi fido valore iniziale

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
### 4) Steady-state and dynamic optimizers

# 4.1) Setpoints

def defSP(t):
    """
    SUMMARY:
    It constructs the setpoints vectors for the steady-state optimisation 
    
    SYNTAX:
    assignment = defSP(t)
  
    ARGUMENTS:
    + t             - Variable that indicates the current time
    
    OUTPUTS:
    + ysp, usp, xsp - Input, output and state setpoint values      
    """ 
    xsp = np.array([-1300.0])           # State setpoints
    SP = 450.0
    if t<= 100:
        ysp = np.array([SP])         # Output setpoint
        usp = np.array([150.])          # Input setpoints
    elif t> 100 and t<= 200:
        ysp = np.array([SP])         # Output setpoint
        usp = np.array([150.])          # Input setpoints     
    elif t> 200 and t<= 300:
        ysp = np.array([SP])         # Output setpoint
        usp = np.array([150.])          # Control setpoints
    else:
        ysp = np.array([SP])         # Output setpoint
        usp = np.array([150.])          # Input setpoints 
        
    return [ysp, usp, xsp]
    
# 4.2) Bounds constraints
## Input bounds
# to be checked ...
umin = -300*np.ones((u.size1(),1))
#umin = 0*np.ones((u.size1(),1))
umax = 300*np.ones((u.size1(),1))

## State bounds
# xmin = np.array([-10.0, -8.0, -10.0])
# xmax = 10.0*np.ones((x.size1(),1))

## Output bounds
ymin = 0.0*np.ones((y.size1(),1))
ymax = 2000.0*np.ones((y.size1(),1))

# 4.3) Steady-state optimization: objective function
Qss = np.diag([100])
Rss = np.zeros((u.size1(),u.size1())) # Control matrix

# 4.4) Dynamic optimization: objective function 
Qy = np.diag([1000])
Q = np.dot(C.T,np.dot(Qy,C))
S = np.diag([100])













